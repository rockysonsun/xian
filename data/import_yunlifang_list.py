#!/usr/bin/env python3
"""
导入云立方机房业务板使用情况 - 列表格式版本
"""

import sqlite3
import openpyxl
from datetime import datetime
import re

DB_PATH = '/Users/rocky/.openclaw/workspace/transmission_resource.db'
EXCEL_PATH = '/Users/rocky/.openclaw/workspace/2026传输资源分配表.xlsx'


def get_or_create_site(cursor, site_name):
    """获取或创建站点"""
    cursor.execute('SELECT id FROM sites WHERE name = ?', (site_name,))
    site = cursor.fetchone()
    if site:
        return site[0]
    else:
        cursor.execute('INSERT INTO sites (name, location) VALUES (?, ?)',
                       (site_name, site_name))
        return cursor.lastrowid


def parse_wavelength(text):
    """从文本中提取波长信息"""
    if not text or text == '空':
        return None
    match = re.search(r'(\d{3}\.\d)', str(text))
    return match.group(1) if match else None


def parse_direction(text):
    """从文本中提取方向信息"""
    if not text or text == '空':
        return None
    text = str(text)
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '腾讯' in text:
        directions.append('腾讯')
    return '-'.join(directions) if directions else None


def import_yunlifang():
    """导入云立方机房数据 - 列表格式"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['云立方机房业务板使用情况']

    # 获取或创建云立方站点
    site_id = get_or_create_site(cursor, '云立方机房')
    print(f"站点ID: {site_id}")

    devices_imported = 0
    cards_imported = 0

    current_device = None
    current_device_id = None
    current_vendor = None
    current_model = None

    # 从第3行开始读取数据 (第1行是标题，第2行是表头)
    for row in range(3, ws.max_row + 1):
        # 读取各列数据
        vendor_cell = ws.cell(row=row, column=1).value  # A列: 品牌
        model_cell = ws.cell(row=row, column=2).value   # B列: 设备型号
        device_code_cell = ws.cell(row=row, column=3).value  # C列: 设备编号
        slot_cell = ws.cell(row=row, column=4).value    # D列: 槽位
        card_type_cell = ws.cell(row=row, column=5).value  # E列: 板卡型号
        usage_cell = ws.cell(row=row, column=6).value   # F列: 使用状况

        # 如果有设备编号，说明是新设备
        if device_code_cell and str(device_code_cell).strip():
            device_code = str(device_code_cell).strip()

            # 确定厂商
            if vendor_cell and str(vendor_cell).strip():
                current_vendor = str(vendor_cell).strip()

            # 确定设备型号
            if model_cell and str(model_cell).strip():
                current_model = str(model_cell).strip()
            else:
                current_model = '未知'

            # 创建设备记录
            cursor.execute('''
                INSERT INTO devices (site_id, name, device_code, vendor, device_type, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (site_id, device_code, device_code, current_vendor or '未知', current_model, 'active', datetime.now()))
            current_device_id = cursor.lastrowid
            devices_imported += 1
            current_device = device_code

            print(f"  导入设备: {device_code} ({current_vendor} {current_model})")

        # 如果有槽位号，导入板卡
        if slot_cell and str(slot_cell).strip().isdigit():
            slot_number = int(str(slot_cell).strip())
            card_type = str(card_type_cell).strip() if card_type_cell else None
            port_info = str(usage_cell).strip() if usage_cell else None

            if card_type and card_type != '空':
                wavelength = parse_wavelength(port_info)
                direction = parse_direction(port_info)
                status = 'empty' if not port_info or port_info == '空' else 'active'

                # 构建描述
                description = None
                if port_info and port_info != '空':
                    parts = []
                    if wavelength:
                        parts.append(f"波长:{wavelength}")
                    if direction:
                        parts.append(f"方向:{direction}")
                    parts.append(f"详情:{port_info}")
                    description = " ".join(parts)

                cursor.execute('''
                    INSERT INTO cards (device_id, slot_number, card_type, status, description)
                    VALUES (?, ?, ?, ?, ?)
                ''', (current_device_id, slot_number, card_type, status, description))
                cards_imported += 1

    conn.commit()
    conn.close()
    wb.close()

    print(f"\n=== 导入完成 ===")
    print(f"设备导入: {devices_imported}")
    print(f"板卡导入: {cards_imported}")


def show_summary():
    """显示导入的数据摘要"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    print("\n=== 云立方机房设备概览 ===")

    # 获取站点ID
    cursor.execute('SELECT id FROM sites WHERE name = ?', ('云立方机房',))
    site = cursor.fetchone()
    if not site:
        print("未找到云立方机房数据")
        return
    site_id = site[0]

    # 按厂商统计
    cursor.execute('''
        SELECT vendor, COUNT(*) as device_count
        FROM devices
        WHERE site_id = ?
        GROUP BY vendor
    ''', (site_id,))

    for row in cursor.fetchall():
        vendor, count = row
        print(f"\n【{vendor}】设备数量: {count}")

        cursor.execute('''
            SELECT name, device_type,
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id) as slot_count,
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id AND status = 'active') as active_count
            FROM devices d
            WHERE site_id = ? AND vendor = ?
        ''', (site_id, vendor))

        for device in cursor.fetchall():
            name, dtype, total, active = device
            print(f"  - {name[:50]:<50} {active}/{total} 槽位在用")

    # 波长统计
    print("\n=== 波长使用情况 ===")
    cursor.execute('''
        SELECT c.description
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ? AND c.status = 'active' AND c.description LIKE '%波长:%'
    ''', (site_id,))

    wavelengths = {}
    for row in cursor.fetchall():
        desc = row[0]
        if desc:
            match = re.search(r'波长:(\d{3}\.\d)', desc)
            if match:
                wl = match.group(1)
                wavelengths[wl] = wavelengths.get(wl, 0) + 1

    sorted_wl = sorted(wavelengths.items(), key=lambda x: float(x[0]))
    print(f"在用波长数量: {len(sorted_wl)}")
    if sorted_wl:
        for i in range(0, len(sorted_wl), 10):
            chunk = sorted_wl[i:i+10]
            print("波长分布: ", ", ".join([f"{w[0]}nm({w[1]})" for w in chunk]))

    # 统计空置槽位
    cursor.execute('''
        SELECT COUNT(*)
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ? AND c.status = 'empty'
    ''', (site_id,))
    empty_count = cursor.fetchone()[0]

    cursor.execute('''
        SELECT COUNT(*)
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ?
    ''', (site_id,))
    total_count = cursor.fetchone()[0]

    if total_count > 0:
        print(f"\n槽位利用率: {total_count - empty_count}/{total_count} ({(total_count - empty_count) / total_count * 100:.1f}%)")

    conn.close()


if __name__ == '__main__':
    print("=" * 60)
    print("云立方机房业务板数据导入工具 (列表格式)")
    print("=" * 60)
    print(f"数据库: {DB_PATH}")
    print(f"Excel文件: {EXCEL_PATH}")

    # 先清空云立方机房的旧数据
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM cards WHERE device_id IN (SELECT id FROM devices WHERE site_id = (SELECT id FROM sites WHERE name = "云立方机房"))')
    cursor.execute('DELETE FROM devices WHERE site_id = (SELECT id FROM sites WHERE name = "云立方机房")')
    conn.commit()
    conn.close()
    print("\n已清空旧数据")

    import_yunlifang()
    show_summary()

    print("\n" + "=" * 60)
    print("✅ 数据导入完成！")
    print("=" * 60)
