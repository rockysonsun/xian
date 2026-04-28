#!/usr/bin/env python3
"""
导入波长连接数据（云立方-常熟等sheet）
"""
import openpyxl
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.models import init_db, get_session, WavelengthConnection


def import_wavelength_sheet(excel_path, sheet_name, db_path='transmission_resource.db'):
    """导入波长连接数据"""
    
    engine = init_db(db_path)
    session = get_session(engine)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        print(f"错误: Sheet '{sheet_name}' 不存在")
        return 0
    
    ws = wb[sheet_name]
    print(f"\n导入: {sheet_name}")
    print(f"总行数: {ws.max_row}")
    
    imported = 0
    current_plane = ""  # 当前平面（处理合并单元格）
    current_a_device = ""  # 当前A端传输设备
    current_z_device = ""  # 当前Z端传输设备（处理合并单元格）
    
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        
        # 读取数据
        plane = row[0].value  # A列: 平面
        local_network_device = row[1].value  # B列: 本端数通设备
        local_network_port = row[2].value  # C列: 本端数通端口号
        local_ip = row[3].value  # D列: 本端接口IP
        otn_tributary_port = row[4].value  # E列: OTN支路侧端口号
        a_transmission_device = row[5].value  # F列: A端传输设备
        otn_line_port_a = row[6].value  # G列: OTN线路侧端口号
        wavelength = row[7].value  # H列: 波长
        otn_line_port_z = row[8].value  # I列: OTN线路侧端口号(Z端)
        z_transmission_device = row[9].value  # J列: Z端传输设备
        otn_tributary_port_z = row[10].value  # K列: Z端OTN支路侧端口号
        remote_ip = row[11].value  # L列: 对端接口IP
        remote_network_port = row[12].value  # M列: 对端数通端口号
        remote_room_device = row[13].value  # N列: 对端机房设备
        remark = row[14].value  # O列: 备注
        other = row[15].value  # P列: 其他
        
        # 更新当前平面（处理合并单元格）
        if plane:
            current_plane = plane
        
        # 更新A端传输设备（处理合并单元格）
        if a_transmission_device:
            current_a_device = a_transmission_device
        
        # 更新Z端传输设备（处理合并单元格）
        if z_transmission_device:
            current_z_device = z_transmission_device
        
        # 跳过空行（没有波长信息或波长为空/空字符串）
        if wavelength is None or str(wavelength).strip() == '' or str(wavelength).strip() == '空':
            continue
        
        # 创建记录
        conn = WavelengthConnection(
            sheet_name=sheet_name,
            plane=current_plane,
            local_network_device=local_network_device if local_network_device else '',
            local_network_port=local_network_port if local_network_port else '',
            local_ip=local_ip if local_ip else '',
            otn_tributary_port=otn_tributary_port if otn_tributary_port else '',
            a_transmission_device=current_a_device,
            otn_line_port_a=otn_line_port_a if otn_line_port_a else '',
            wavelength=str(wavelength).strip() if wavelength else '',
            otn_line_port_z=otn_line_port_z if otn_line_port_z else '',
            z_transmission_device=current_z_device,
            otn_tributary_port_z=otn_tributary_port_z if otn_tributary_port_z else '',
            remote_ip=remote_ip if remote_ip else '',
            remote_network_port=remote_network_port if remote_network_port else '',
            remote_room_device=remote_room_device if remote_room_device else '',
            remark=remark if remark else '',
            other=other if other else ''
        )
        
        session.add(conn)
        imported += 1
        
        if imported % 10 == 0:
            print(f"  已导入 {imported} 条...")
    
    session.commit()
    print(f"  完成: 导入 {imported} 条记录")
    session.close()
    return imported


if __name__ == '__main__':
    excel_file = '2026传输资源分配表.xlsx'
    if not os.path.exists(excel_file):
        print(f"错误: 找不到文件 {excel_file}")
        sys.exit(1)
    
    # 导入云立方-常熟
    import_wavelength_sheet(excel_file, '云立方-常熟')
