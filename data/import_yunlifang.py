#!/usr/bin/env python3
"""
导入云立方机房业务板使用情况到数据库
"""

import sqlite3
import openpyxl
from datetime import datetime
import re

DB_PATH = '/Users/rocky/.openclaw/workspace/transmission_resource.db'
EXCEL_PATH = '/Users/rocky/.openclaw/workspace/2026传输资源分配表.xlsx'

def init_database():
    """初始化数据库表结构"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 机房表
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS rooms (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        location TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # 设备表
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS devices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        room_id INTEGER,
        device_name TEXT NOT NULL,
        device_model TEXT,
        device_type TEXT,
        slot_count INTEGER,
        plane TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (room_id) REFERENCES rooms(id)
    )
    ''')
    
    # 检查并添加 room_id 列（如果不存在）
    cursor.execute("PRAGMA table_info(devices)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'room_id' not in columns:
        cursor.execute('ALTER TABLE devices ADD COLUMN room_id INTEGER')
        print("已添加 room_id 列到 devices 表")
    
    # 板卡表
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS cards (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id INTEGER,
        slot_number INTEGER,
        card_type TEXT,
        port_info TEXT,
        wavelength TEXT,
        direction TEXT,
        status TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (device_id) REFERENCES devices(id)
    )
    ''')
    
    conn.commit()
    conn.close()
    print("数据库表结构初始化完成")

def parse_wavelength(text):
    """从文本中提取波长信息"""
    if not text or text == '空':
        return None
    # 匹配 xxx.x 格式的波长
    match = re.search(r'(\d{3}\.\d)', text)
    return match.group(1) if match else None

def parse_direction(text):
    """从文本中提取方向信息"""
    if not text or text == '空':
        return None
    # 提取方向关键词
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '腾讯' in text:
        directions.append('腾讯')
    return '-'.join(directions) if directions else None

def import_yunlifang():
    """导入云立方机房数据"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['云立方机房业务板使用情况']
    
    # 获取或创建云立方机房
    cursor.execute('SELECT id FROM rooms WHERE name = ?', ('云立方机房',))
    room = cursor.fetchone()
    if room:
        room_id = room[0]
    else:
        cursor.execute('INSERT INTO rooms (name, location) VALUES (?, ?)', 
                      ('云立方机房', '云立方'))
        room_id = cursor.lastrowid
    
    print(f"机房ID: {room_id}")
    
    # 处理华为平面数据 (第3-10行)
    print("\n=== 导入华为平面数据 ===")
    huawei_devices = []
    
    # 读取设备信息 (第3行)
    row_idx = 3
    col = 2  # 从B列开始
    
    while col <= ws.max_column:
        device_cell = ws.cell(row=row_idx, column=col)
        if device_cell.value and 'DC908' in str(device_cell.value):
            device_name = str(device_cell.value).strip()
            # 提取设备型号
            if 'DC908Pro' in device_name:
                device_model = 'DC908Pro'
            else:
                device_model = 'DC908'
            
            huawei_devices.append({
                'name': device_name,
                'model': device_model,
                'col': col
            })
            print(f"发现设备: {device_name}")
        col += 4  # 每4列一个设备
    
    # 导入每个设备的板卡信息
    for device_info in huawei_devices:
        # 创建设备记录
        cursor.execute('''
            INSERT INTO devices (room_id, device_name, device_model, device_type, plane)
            VALUES (?, ?, ?, ?, ?)
        ''', (room_id, device_info['name'], device_info['model'], '传输设备', '华为'))
        device_id = cursor.lastrowid
        
        # 读取该设备的所有槽位 (第4-11行)
        for slot_row in range(4, 12):
            slot_num_cell = ws.cell(row=slot_row, column=device_info['col'])
            card_type_cell = ws.cell(row=slot_row, column=device_info['col'] + 1)
            port_info_cell = ws.cell(row=slot_row, column=device_info['col'] + 2)
            
            slot_number = slot_num_cell.value
            card_type = card_type_cell.value
            port_info = port_info_cell.value
            
            if slot_number and str(slot_number).isdigit():
                slot_number = int(slot_number)
                wavelength = parse_wavelength(str(port_info) if port_info else '')
                direction = parse_direction(str(port_info) if port_info else '')
                status = '空置' if not card_type or card_type == '空' else '在用'
                
                cursor.execute('''
                    INSERT INTO cards (device_id, slot_number, card_type, port_info, wavelength, direction, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (device_id, slot_number, card_type, port_info, wavelength, direction, status))
    
    # 处理光迅平面数据 (从第13行开始)
    print("\n=== 导入光迅平面数据 ===")
    
    row = 13
    while row <= ws.max_row:
        device_cell = ws.cell(row=row, column=2)
        if device_cell.value and ('O2' in str(device_cell.value) or 'T1200' in str(device_cell.value)):
            device_name = str(device_cell.value).strip()
            
            # 提取设备型号
            if 'O2' in device_name:
                device_model = 'O2'
            elif 'T1200' in device_name:
                device_model = 'T1200'
            else:
                device_model = '未知'
            
            # 创建设备记录
            cursor.execute('''
                INSERT INTO devices (room_id, device_name, device_model, device_type, plane)
                VALUES (?, ?, ?, ?, ?)
            ''', (room_id, device_name, device_model, '光层设备', '光迅'))
            device_id = cursor.lastrowid
            
            # 读取该设备的所有槽位 (当前行开始，最多8行)
            for slot_offset in range(8):
                slot_row = row + slot_offset
                if slot_row > ws.max_row:
                    break
                    
                slot_num_cell = ws.cell(row=slot_row, column=3)
                card_type_cell = ws.cell(row=slot_row, column=4)
                port_info_cell = ws.cell(row=slot_row, column=5)
                
                slot_number = slot_num_cell.value
                card_type = card_type_cell.value
                port_info = port_info_cell.value
                
                if slot_number and str(slot_number).isdigit():
                    slot_number = int(slot_number)
                    wavelength = parse_wavelength(str(port_info) if port_info else '')
                    direction = parse_direction(str(port_info) if port_info else '')
                    status = '空置' if not card_type or card_type == '空' else '在用'
                    
                    cursor.execute('''
                        INSERT INTO cards (device_id, slot_number, card_type, port_info, wavelength, direction, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (device_id, slot_number, card_type, port_info, wavelength, direction, status))
            
            # 跳过已处理的行
            row += 8
        else:
            row += 1
    
    conn.commit()
    
    # 统计导入结果
    cursor.execute('SELECT COUNT(*) FROM devices WHERE room_id = ?', (room_id,))
    device_count = cursor.fetchone()[0]
    
    cursor.execute('''
        SELECT COUNT(*) FROM cards c 
        JOIN devices d ON c.device_id = d.id 
        WHERE d.room_id = ?
    ''', (room_id,))
    card_count = cursor.fetchone()[0]
    
    cursor.execute('''
        SELECT COUNT(*) FROM cards c 
        JOIN devices d ON c.device_id = d.id 
        WHERE d.room_id = ? AND c.status = '在用'
    ''', (room_id,))
    active_card_count = cursor.fetchone()[0]
    
    conn.close()
    wb.close()
    
    print(f"\n=== 导入完成 ===")
    print(f"设备总数: {device_count}")
    print(f"板卡总数: {card_count}")
    print(f"在用板卡: {active_card_count}")
    print(f"空置槽位: {card_count - active_card_count}")

def show_summary():
    """显示导入的数据摘要"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("\n=== 云立方机房设备概览 ===")
    
    # 按平面统计
    cursor.execute('''
        SELECT plane, COUNT(*) as device_count 
        FROM devices d
        JOIN rooms r ON d.room_id = r.id
        WHERE r.name = '云立方机房'
        GROUP BY plane
    ''')
    for row in cursor.fetchall():
        print(f"\n【{row[0]}平面】设备数量: {row[1]}")
        
        # 显示该平面的设备列表
        cursor.execute('''
            SELECT device_name, device_model, 
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id) as slot_count,
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id AND status = '在用') as active_count
            FROM devices d
            JOIN rooms r ON d.room_id = r.id
            WHERE r.name = '云立方机房' AND d.plane = ?
        ''', (row[0],))
        
        for device in cursor.fetchall():
            print(f"  - {device[0]} ({device[1]}): {device[3]}/{device[2]} 槽位在用")
    
    # 波长统计
    print("\n=== 波长使用情况 ===")
    cursor.execute('''
        SELECT wavelength, COUNT(*) as count
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        JOIN rooms r ON d.room_id = r.id
        WHERE r.name = '云立方机房' AND c.wavelength IS NOT NULL
        GROUP BY wavelength
        ORDER BY CAST(wavelength AS REAL)
    ''')
    
    wavelengths = cursor.fetchall()
    print(f"在用波长数量: {len(wavelengths)}")
    print("波长列表: ", ", ".join([f"{w[0]}nm({w[1]})" for w in wavelengths[:10]]), 
          "..." if len(wavelengths) > 10 else "")
    
    conn.close()

if __name__ == '__main__':
    print("开始导入云立方机房业务板数据...")
    print(f"数据库: {DB_PATH}")
    print(f"Excel文件: {EXCEL_PATH}")
    
    init_database()
    import_yunlifang()
    show_summary()
    
    print("\n✅ 数据导入完成！")
