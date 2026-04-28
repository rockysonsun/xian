#!/usr/bin/env python3
"""
导入云立方机房业务板使用情况 - V5版本
优化点：
1. 正确处理 N2MD02 的 IN1OUT1/IN2OUT2 双波长
2. 完整保留原始"使用状况"作为 description
"""

import sqlite3
import openpyxl
from datetime import datetime
import re

DB_PATH = '/Users/rocky/.openclaw/workspace/transmission_resource.db'
EXCEL_PATH = '/Users/rocky/.openclaw/workspace/2026传输资源分配表.xlsx'


def get_or_create_site(cursor, site_name):
    """获取或创建站点"""
    cursor.execute('SELECT id FROM sites WHERE name = ?', (site_name,))
    site = cursor.fetchone()
    if site:
        return site[0]
    else:
        cursor.execute('INSERT INTO sites (name, code, location) VALUES (?, ?, ?)', 
                      (site_name, site_name[:3].upper(), site_name))
        return cursor.lastrowid


def parse_n2md02_wavelengths(usage_text):
    """
    专门解析 N2MD02 的双波长格式
    支持格式：
    - IN1/OUT1(嘉定方向192.1)/IN2/OUT2(嘉定方向192.2)
    - IN1OUT1-云立方-吴江-192.2/IN2OUT2-空置
    - IN1OUT1-空置/IN2OUT2-云立方-嘉定-常熟-FOADM-193.7
    
    返回: [(port, wavelength, direction), ...]
    """
    if not usage_text or usage_text == '空':
        return []
    
    text = str(usage_text)
    results = []
    
    # 模式1: IN1/OUT1(方向波长)/IN2/OUT2(方向波长)
    # 例如: IN1/OUT1(嘉定方向192.1)/IN2/OUT2(嘉定方向192.2)
    pattern1 = r'IN1[/]?OUT1\(([^)]+)\)(?:.*?IN2[/]?OUT2\(([^)]+)\))?'
    match1 = re.search(pattern1, text)
    if match1:
        # IN1OUT1
        in1_content = match1.group(1)
        wl_match = re.search(r'(\d{3}\.\d)', in1_content)
        dir_match = re.search(r'(.+?方向)', in1_content)
        if wl_match:
            results.append({
                'port': 'IN1OUT1',
                'wavelength': wl_match.group(1),
                'direction': dir_match.group(1) if dir_match else None
            })
        
        # IN2OUT2
        if match1.group(2):
            in2_content = match1.group(2)
            wl_match = re.search(r'(\d{3}\.\d)', in2_content)
            dir_match = re.search(r'(.+?方向)', in2_content)
            if wl_match:
                results.append({
                    'port': 'IN2OUT2',
                    'wavelength': wl_match.group(1),
                    'direction': dir_match.group(1) if dir_match else None
                })
        return results
    
    # 模式2: IN1OUT1-xxx-波长/IN2OUT2-xxx-波长
    # 提取所有 IN1OUT1 和 IN2OUT2 的波长
    in1_pattern = r'IN1[/]?OUT1[^\d]*(\d{3}\.\d)'
    in2_pattern = r'IN2[/]?OUT2[^\d]*(\d{3}\.\d)'
    
    in1_match = re.search(in1_pattern, text)
    in2_match = re.search(in2_pattern, text)
    
    # 提取方向
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    direction = '-'.join(directions) if directions else None
    
    if in1_match:
        results.append({
            'port': 'IN1OUT1',
            'wavelength': in1_match.group(1),
            'direction': direction
        })
    
    if in2_match:
        results.append({
            'port': 'IN2OUT2',
            'wavelength': in2_match.group(1),
            'direction': direction
        })
    
    return results


def parse_single_wavelength(usage_text):
    """解析单波长板卡的波长"""
    if not usage_text or usage_text == '空':
        return None
    
    text = str(usage_text)
    
    # 提取波长
    wl_match = re.search(r'(\d{3}\.\d)', text)
    if not wl_match:
        return None
    
    # 提取方向
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '腾讯' in text:
        directions.append('腾讯')
    
    return {
        'port': 'IN/OUT',
        'wavelength': wl_match.group(1),
        'direction': '-'.join(directions) if directions else None
    }


def import_yunlifang():
    """导入云立方机房数据"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['云立方机房业务板使用情况']
    
    # 获取或创建云立方站点
    site_id = get_or_create_site(cursor, '云立方机房')
    print(f"站点ID: {site_id}")
    
    devices_imported = 0
    cards_imported = 0
    ports_imported = 0
    
    current_device_id = None
    current_vendor = None
    current_model = None
    
    # 从第3行开始读取数据
    for row in range(3, ws.max_row + 1):
        # 读取各列数据
        vendor_cell = ws.cell(row=row, column=1).value  # A列: 品牌
        model_cell = ws.cell(row=row, column=2).value   # B列: 设备型号
        device_code_cell = ws.cell(row=row, column=3).value  # C列: 设备编号
        slot_cell = ws.cell(row=row, column=4).value    # D列: 槽位
        card_type_cell = ws.cell(row=row, column=5).value  # E列: 板卡型号
        usage_cell = ws.cell(row=row, column=6).value   # F列: 使用状况
        
        # 跳过表头行
        if device_code_cell == '设备编号':
            continue
        
        # 如果有设备编号，说明是新设备
        if device_code_cell and str(device_code_cell).strip() and str(device_code_cell).strip() != 'None':
            device_code = str(device_code_cell).strip()
            
            # 确定厂商
            if vendor_cell and str(vendor_cell).strip():
                current_vendor = str(vendor_cell).strip()
            
            # 确定设备型号
            if model_cell and str(model_cell).strip():
                current_model = str(model_cell).strip()
            else:
                current_model = '未知'
            
            # 创建设备记录
            cursor.execute('''
                INSERT INTO devices (site_id, name, device_code, vendor, device_type, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (site_id, device_code, device_code, current_vendor or '未知', current_model, 'active', datetime.now()))
            current_device_id = cursor.lastrowid
            devices_imported += 1
            
            print(f"  导入设备: {device_code} ({current_vendor} {current_model})")
        
        # 如果有槽位号，导入板卡
        if slot_cell and str(slot_cell).strip().isdigit() and current_device_id:
            slot_number = int(str(slot_cell).strip())
            card_type = str(card_type_cell).strip() if card_type_cell else None
            usage = str(usage_cell).strip() if usage_cell else None
            
            if card_type and card_type != '空':
                # 完整保留原始使用状况作为 description
                description = usage if usage and usage != '空' else None
                
                # 解析波长
                wavelengths = []
                if card_type.upper() == 'N2MD02':
                    wavelengths = parse_n2md02_wavelengths(usage)
                else:
                    wl = parse_single_wavelength(usage)
                    if wl:
                        wavelengths = [wl]
                
                status = 'empty' if not usage or usage == '空' else 'active'
                
                # 创建板卡记录
                cursor.execute('''
                    INSERT INTO cards (device_id, slot_number, card_type, status, description)
                    VALUES (?, ?, ?, ?, ?)
                ''', (current_device_id, slot_number, card_type, status, description))
                card_id = cursor.lastrowid
                cards_imported += 1
                
                # 创建端口记录
                for wl_info in wavelengths:
                    cursor.execute('''
                        INSERT INTO ports (card_id, port_number, port_type, status, wavelength, description)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (card_id, wl_info['port'], '光口', 'active', float(wl_info['wavelength']), wl_info.get('direction')))
                    ports_imported += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"\n=== 导入完成 ===")
    print(f"设备导入: {devices_imported}")
    print(f"板卡导入: {cards_imported}")
    print(f"波长端口导入: {ports_imported}")


def verify_n2md02():
    """验证 N2MD02 的双波长解析是否正确"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("\n=== N2MD02 双波长验证 ===")
    
    cursor.execute('''
        SELECT 
            d.name as device_name,
            c.slot_number,
            c.card_type,
            c.description as original_usage,
            GROUP_CONCAT(p.port_number || ':' || p.wavelength, ' | ') as parsed_wavelengths
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        LEFT JOIN ports p ON c.id = p.card_id
        WHERE d.site_id = (SELECT id FROM sites WHERE name = '云立方机房')
          AND c.card_type = 'N2MD02'
        GROUP BY d.name, c.slot_number
        ORDER BY d.name, c.slot_number
    ''')
    
    print("\n设备 | 槽位 | 原始使用状况 | 解析出的波长")
    print("-" * 80)
    for row in cursor.fetchall():
        device, slot, card, usage, wl = row
        print(f"{device} | 槽位{slot} | {usage[:40]}... | {wl}")
    
    conn.close()


if __name__ == '__main__':
    print("=" * 70)
    print("云立方机房业务板数据导入工具 (V5 - 优化双波长解析)")
    print("=" * 70)
    print("\n优化点:")
    print("  1. 正确解析 N2MD02 的 IN1OUT1/IN2OUT2 双波长")
    print("  2. 完整保留原始'使用状况'字段")
    
    # 先清空云立方机房的旧数据
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM ports WHERE card_id IN (SELECT id FROM cards WHERE device_id IN (SELECT id FROM devices WHERE site_id = (SELECT id FROM sites WHERE name = "云立方机房")))')
    cursor.execute('DELETE FROM cards WHERE device_id IN (SELECT id FROM devices WHERE site_id = (SELECT id FROM sites WHERE name = "云立方机房"))')
    cursor.execute('DELETE FROM devices WHERE site_id = (SELECT id FROM sites WHERE name = "云立方机房")')
    conn.commit()
    conn.close()
    print("\n已清空云立方机房旧数据")
    
    import_yunlifang()
    verify_n2md02()
    
    print("\n" + "=" * 70)
    print("✅ 数据导入完成！")
    print("=" * 70)
