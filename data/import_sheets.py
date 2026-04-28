#!/usr/bin/env python3
"""
导入Excel中的4个机房业务板使用情况sheet
"""
import openpyxl
import sys
import os

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.models import init_db, get_session, BusinessBoard

def import_business_boards(excel_path, db_path='transmission_resource.db'):
    """导入业务板数据"""
    
    # 初始化数据库
    engine = init_db(db_path)
    session = get_session(engine)
    
    # 要导入的sheet
    sheets_to_import = [
        '云立方机房业务板使用情况',
        '嘉定机房业务板使用情况', 
        '常熟机房业务板使用情况',
        '吴江机房业务板使用情况'
    ]
    
    # 打开Excel
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    total_imported = 0
    
    for sheet_name in sheets_to_import:
        if sheet_name not in wb.sheetnames:
            print(f"警告: Sheet '{sheet_name}' 不存在，跳过")
            continue
            
        ws = wb[sheet_name]
        room_name = ws.cell(row=1, column=2).value  # 机房名称在B1单元格
        
        print(f"\n导入: {sheet_name} (机房: {room_name})")
        
        # 从第4行开始读取数据（第3行是表头）
        imported = 0
        current_device = {}  # 当前设备信息（处理合并单元格）
        
        for row_idx in range(4, ws.max_row + 1):
            row = ws[row_idx]
            
            # 读取各列数据
            device_brand = row[0].value  # A列: 设备品牌
            device_model = row[1].value  # B列: 设备型号
            device_code = row[2].value   # C列: 设备编号
            slot = row[3].value          # D列: 槽位
            board_model = row[4].value   # E列: 板卡型号
            usage_status = row[5].value  # F列: 使用状况
            
            # 如果设备信息不为空，更新当前设备
            if device_brand:
                current_device['brand'] = device_brand
            if device_model:
                current_device['model'] = device_model
            if device_code:
                current_device['code'] = device_code
            
            # 跳过空行（没有槽位信息）
            if slot is None:
                continue
                
            # 尝试转换槽位为整数
            try:
                slot_num = int(slot)
            except (ValueError, TypeError):
                continue
            
            # 创建记录
            board = BusinessBoard(
                room_name=room_name,
                device_brand=current_device.get('brand', ''),
                device_model=current_device.get('model', ''),
                device_code=current_device.get('code', ''),
                slot=slot_num,
                board_model=board_model if board_model else '',
                usage_status=usage_status if usage_status else ''
            )
            
            session.add(board)
            imported += 1
            
            if imported % 50 == 0:
                print(f"  已导入 {imported} 条...")
        
        session.commit()
        print(f"  完成: 导入 {imported} 条记录")
        total_imported += imported
    
    session.close()
    print(f"\n✅ 总计导入: {total_imported} 条记录")
    return total_imported

if __name__ == '__main__':
    excel_file = '2026传输资源分配表.xlsx'
    if not os.path.exists(excel_file):
        print(f"错误: 找不到文件 {excel_file}")
        sys.exit(1)
    
    import_business_boards(excel_file)
