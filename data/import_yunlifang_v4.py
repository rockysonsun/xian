#!/usr/bin/env python3
"""
导入云立方机房业务板使用情况 - V4版本
支持N2MD02板卡的双波长逻辑 (IN1OUT1 + IN2OUT2)
"""

import sqlite3
import openpyxl
from datetime import datetime
import re

DB_PATH = '/Users/rocky/.openclaw/workspace/transmission_resource.db'
EXCEL_PATH = '/Users/rocky/.openclaw/workspace/2026传输资源分配表.xlsx'


def get_or_create_site(cursor, site_name):
    """获取或创建站点"""
    cursor.execute('SELECT id FROM sites WHERE name = ?', (site_name,))
    site = cursor.fetchone()
    if site:
        return site[0]
    else:
        cursor.execute('INSERT INTO sites (name, code, location) VALUES (?, ?, ?)', 
                      (site_name, site_name[:3].upper(), site_name))
        return cursor.lastrowid


def parse_wavelengths(text, card_type):
    """
    从文本中提取波长信息
    N2MD02板卡支持双波长：IN1OUT1 和 IN2OUT2
    """
    if not text or text == '空':
        return []
    
    wavelengths = []
    text = str(text)
    
    # N2MD02 双波长板卡 - 分别提取 IN1OUT1 和 IN2OUT2 的波长
    if card_type and 'N2MD02' in str(card_type).upper():
        # 尝试匹配 IN1OUT1-xxx-波长/IN2OUT2-xxx-波长 格式
        # 例如: IN1OUT1-云立方-吴江-192.2/IN2OUT2-空置
        # 例如: IN1OUT1-云立方-常熟-直达-192.7/IN2OUT2-空置
        
        # 提取所有波长
        wl_matches = re.findall(r'(\d{3}\.\d)', text)
        
        # 判断是 IN1OUT1 还是 IN2OUT2 的波长
        if 'IN1OUT1' in text or 'IN1/OUT1' in text:
            for wl in wl_matches:
                wavelengths.append({'port': 'IN1OUT1', 'wl': wl})
        
        if 'IN2OUT2' in text or 'IN2/OUT2' in text:
            for wl in wl_matches:
                # 避免重复添加同一个波长
                if not any(w['wl'] == wl for w in wavelengths):
                    wavelengths.append({'port': 'IN2OUT2', 'wl': wl})
        
        # 如果没有明确标识，但有两个波长，默认分配
        if len(wl_matches) == 2 and len(wavelengths) == 0:
            wavelengths = [
                {'port': 'IN1OUT1', 'wl': wl_matches[0]},
                {'port': 'IN2OUT2', 'wl': wl_matches[1]}
            ]
        elif len(wl_matches) == 1 and len(wavelengths) == 0:
            wavelengths = [{'port': 'IN1OUT1', 'wl': wl_matches[0]}]
    
    else:
        # 单波长板卡 - 只提取第一个波长
        wl_match = re.search(r'(\d{3}\.\d)', text)
        if wl_match:
            wavelengths = [{'port': 'IN/OUT', 'wl': wl_match.group(1)}]
    
    return wavelengths


def parse_direction(text):
    """从文本中提取方向信息"""
    if not text or text == '空':
        return None
    text = str(text)
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '腾讯' in text:
        directions.append('腾讯')
    return '-'.join(directions) if directions else None


def get_wavelength_count(card_type):
    """根据板卡类型返回波长数量"""
    if not card_type:
        return 0
    card_type = str(card_type).upper()
    if 'N2MD02' in card_type:
        return 2
    elif 'N2MS04' in card_type:
        return 1
    elif 'N2MD04CE' in card_type:
        return 2
    elif 'T1X200S' in card_type:
        return 1
    elif 'T04X1' in card_type:
        return 1
    return 0


def import_yunlifang():
    """导入云立方机房数据"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['云立方机房业务板使用情况']
    
    # 获取或创建云立方站点
    site_id = get_or_create_site(cursor, '云立方机房')
    print(f"站点ID: {site_id}")
    
    devices_imported = 0
    cards_imported = 0
    wavelengths_imported = 0
    
    current_device = None
    current_device_id = None
    current_vendor = None
    current_model = None
    
    # 从第3行开始读取数据 (第1行标题，第2行表头)
    for row in range(3, ws.max_row + 1):
        # 读取各列数据
        vendor_cell = ws.cell(row=row, column=1).value  # A列: 品牌
        model_cell = ws.cell(row=row, column=2).value   # B列: 设备型号
        device_code_cell = ws.cell(row=row, column=3).value  # C列: 设备编号
        slot_cell = ws.cell(row=row, column=4).value    # D列: 槽位
        card_type_cell = ws.cell(row=row, column=5).value  # E列: 板卡型号
        usage_cell = ws.cell(row=row, column=6).value   # F列: 使用状况
        
        # 跳过表头行
        if device_code_cell == '设备编号':
            continue
        
        # 如果有设备编号，说明是新设备
        if device_code_cell and str(device_code_cell).strip() and str(device_code_cell).strip() != 'None':
            device_code = str(device_code_cell).strip()
            
            # 确定厂商
            if vendor_cell and str(vendor_cell).strip():
                current_vendor = str(vendor_cell).strip()
            
            # 确定设备型号
            if model_cell and str(model_cell).strip():
                current_model = str(model_cell).strip()
            else:
                current_model = '未知'
            
            # 创建设备记录
            cursor.execute('''
                INSERT INTO devices (site_id, name, device_code, vendor, device_type, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (site_id, device_code, device_code, current_vendor or '未知', current_model, 'active', datetime.now()))
            current_device_id = cursor.lastrowid
            devices_imported += 1
            current_device = device_code
            
            print(f"  导入设备: {device_code} ({current_vendor} {current_model})")
        
        # 如果有槽位号，导入板卡
        if slot_cell and str(slot_cell).strip().isdigit() and current_device_id:
            slot_number = int(str(slot_cell).strip())
            card_type = str(card_type_cell).strip() if card_type_cell else None
            port_info = str(usage_cell).strip() if usage_cell else None
            
            if card_type and card_type != '空':
                # 解析波长（支持双波长）
                wavelengths = parse_wavelengths(port_info, card_type)
                wl_count = get_wavelength_count(card_type)
                direction = parse_direction(port_info)
                status = 'empty' if not port_info or port_info == '空' else 'active'
                
                # 构建描述信息
                description_parts = []
                if wavelengths:
                    for wl in wavelengths:
                        description_parts.append(f"{wl['port']}: {wl['wl']}nm")
                if direction:
                    description_parts.append(f"方向: {direction}")
                if port_info and port_info != '空':
                    description_parts.append(f"详情: {port_info[:50]}")
                
                description = " | ".join(description_parts) if description_parts else None
                
                # 创建板卡记录
                cursor.execute('''
                    INSERT INTO cards (device_id, slot_number, card_type, status, description)
                    VALUES (?, ?, ?, ?, ?)
                ''', (current_device_id, slot_number, card_type, status, description))
                card_id = cursor.lastrowid
                cards_imported += 1
                
                # 如果有波长，创建端口记录
                for wl_info in wavelengths:
                    cursor.execute('''
                        INSERT INTO ports (card_id, port_number, port_type, status, wavelength)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (card_id, wl_info['port'], '光口', 'active', float(wl_info['wl'])))
                    wavelengths_imported += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"\n=== 导入完成 ===")
    print(f"设备导入: {devices_imported}")
    print(f"板卡导入: {cards_imported}")
    print(f"波长端口导入: {wavelengths_imported}")


def show_summary():
    """显示导入的数据摘要"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("\n=== 云立方机房设备概览 ===")
    
    # 获取站点ID
    cursor.execute('SELECT id FROM sites WHERE name = ?', ('云立方机房',))
    site = cursor.fetchone()
    if not site:
        print("未找到云立方机房数据")
        return
    site_id = site[0]
    
    # 按厂商统计
    cursor.execute('''
        SELECT vendor, COUNT(*) as device_count 
        FROM devices
        WHERE site_id = ?
        GROUP BY vendor
    ''', (site_id,))
    
    for row in cursor.fetchall():
        vendor, count = row
        print(f"\n【{vendor}】设备数量: {count}")
        
        cursor.execute('''
            SELECT name, device_type, 
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id) as slot_count,
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id AND status = 'active') as active_count
            FROM devices d
            WHERE site_id = ? AND vendor = ?
        ''', (site_id, vendor))
        
        for device in cursor.fetchall():
            name, dtype, total, active = device
            print(f"  - {name[:50]:<50} {active}/{total} 槽位在用")
    
    # 波长统计（按板卡类型）
    print("\n=== 波长分布（按板卡类型）===")
    cursor.execute('''
        SELECT c.card_type, COUNT(p.id) as wl_count
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        LEFT JOIN ports p ON c.id = p.card_id
        WHERE d.site_id = ? AND c.status = 'active'
        GROUP BY c.card_type
    ''', (site_id,))
    
    for row in cursor.fetchall():
        card_type, wl_count = row
        wl_capacity = get_wavelength_count(card_type)
        print(f"  {card_type}: {wl_count} 个波长 (容量: {wl_capacity})")
    
    # N2MD02 双波长详情
    print("\n=== N2MD02 双波长板卡详情 ===")
    cursor.execute('''
        SELECT d.name, c.slot_number, p.port_number, p.wavelength
        FROM ports p
        JOIN cards c ON p.card_id = c.id
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ? AND c.card_type LIKE '%N2MD02%'
        ORDER BY d.name, c.slot_number, p.port_number
    ''', (site_id,))
    
    current_device = None
    current_slot = None
    for row in cursor.fetchall():
        device, slot, port, wl = row
        if device != current_device or slot != current_slot:
            print(f"\n  {device} - 槽位{slot}:")
            current_device = device
            current_slot = slot
        print(f"    {port}: {wl}nm")
    
    conn.close()


if __name__ == '__main__':
    print("=" * 60)
    print("云立方机房业务板数据导入工具 (V4 - 双波长支持)")
    print("=" * 60)
    print(f"数据库: {DB_PATH}")
    print(f"Excel文件: {EXCEL_PATH}")
    print("\n支持N2MD02板卡的双波长逻辑:")
    print("  - IN1OUT1: 第一个波长")
    print("  - IN2OUT2: 第二个波长")
    
    import_yunlifang()
    show_summary()
    
    print("\n" + "=" * 60)
    print("✅ 数据导入完成！")
    print("=" * 60)
