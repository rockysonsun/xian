#!/usr/bin/env python3
"""
导入云立方机房业务板使用情况到数据库 (最终版)
"""

import sqlite3
import openpyxl
from datetime import datetime
import re

DB_PATH = '/Users/rocky/.openclaw/workspace/transmission_resource.db'
EXCEL_PATH = '/Users/rocky/.openclaw/workspace/2026传输资源分配表.xlsx'

def get_or_create_site(cursor, site_name):
    """获取或创建站点"""
    cursor.execute('SELECT id FROM sites WHERE name = ?', (site_name,))
    site = cursor.fetchone()
    if site:
        return site[0]
    else:
        cursor.execute('INSERT INTO sites (name, location) VALUES (?, ?)', 
                      (site_name, site_name))
        return cursor.lastrowid

def parse_wavelength(text):
    """从文本中提取波长信息"""
    if not text or text == '空':
        return None
    match = re.search(r'(\d{3}\.\d)', str(text))
    return match.group(1) if match else None

def parse_direction(text):
    """从文本中提取方向信息"""
    if not text or text == '空':
        return None
    text = str(text)
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '腾讯' in text:
        directions.append('腾讯')
    return '-'.join(directions) if directions else None

def import_yunlifang():
    """导入云立方机房数据"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['云立方机房业务板使用情况']
    
    # 获取或创建云立方站点
    site_id = get_or_create_site(cursor, '云立方机房')
    print(f"站点ID: {site_id}")
    
    devices_imported = 0
    cards_imported = 0
    
    # ========== 处理华为平面数据 ==========
    print("\n=== 导入华为平面数据 ===")
    
    # 第2行是设备名称行
    device_row = 2
    col = 2  # 从B列开始
    
    huawei_devices = []
    
    # 先收集所有华为设备
    while col <= ws.max_column:
        device_cell = ws.cell(row=device_row, column=col)
        if device_cell.value and 'DC908' in str(device_cell.value):
            device_name = str(device_cell.value).strip()
            
            # 提取设备型号
            if 'DC908Pro' in device_name:
                device_model = 'DC908Pro'
            else:
                device_model = 'DC908'
            
            # 提取设备编码
            match = re.search(r'(\d+-\d+)', device_name)
            device_code = match.group(1) if match else device_name
            
            huawei_devices.append({
                'name': device_name,
                'code': device_code,
                'model': device_model,
                'col': col
            })
            print(f"  发现设备: {device_name}")
        col += 3  # 每3列一个设备 (设备名、槽位号、板卡类型、端口信息)
    
    # 导入每个设备的板卡
    for device_info in huawei_devices:
        # 创建设备记录
        cursor.execute('''
            INSERT INTO devices (site_id, name, device_code, vendor, device_type, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (site_id, device_info['name'], device_info['code'], 'HUAWEI', device_info['model'], 'active', datetime.now()))
        device_id = cursor.lastrowid
        devices_imported += 1
        
        # 读取槽位信息 (第3-10行，对应槽位1-8)
        for slot_num in range(1, 9):
            slot_row = 2 + slot_num  # 第3行对应槽位1
            
            slot_num_cell = ws.cell(row=slot_row, column=device_info['col'])
            card_type_cell = ws.cell(row=slot_row, column=device_info['col'] + 1)
            port_info_cell = ws.cell(row=slot_row, column=device_info['col'] + 2)
            
            # 验证槽位号
            actual_slot = slot_num_cell.value
            if actual_slot != slot_num:
                continue
            
            card_type = card_type_cell.value
            port_info = port_info_cell.value
            
            wavelength = parse_wavelength(port_info)
            direction = parse_direction(port_info)
            status = 'empty' if not card_type or card_type == '空' else 'active'
            
            # 构建描述
            description = None
            if port_info and port_info != '空':
                parts = []
                if wavelength:
                    parts.append(f"波长:{wavelength}")
                if direction:
                    parts.append(f"方向:{direction}")
                parts.append(f"详情:{port_info}")
                description = " ".join(parts)
            
            cursor.execute('''
                INSERT INTO cards (device_id, slot_number, card_type, status, description)
                VALUES (?, ?, ?, ?, ?)
            ''', (device_id, slot_num, card_type, status, description))
            cards_imported += 1
    
    # ========== 处理光迅平面数据 ==========
    print("\n=== 导入光迅平面数据 ===")
    
    # 找到"光迅"行 (第11行)
    guangxun_start_row = 11
    row = guangxun_start_row
    
    while row <= ws.max_row:
        # 检查是否是设备行
        col2_value = ws.cell(row=row, column=2).value
        
        if col2_value and ('O2' in str(col2_value) or 'T1200' in str(col2_value)):
            device_name = str(col2_value).strip()
            
            # 提取设备型号
            if 'O2' in device_name:
                device_model = 'O2'
            elif 'T1200' in device_name:
                device_model = 'T1200'
            else:
                device_model = '未知'
            
            device_code = f"GX-{row}"
            
            # 创建设备
            cursor.execute('''
                INSERT INTO devices (site_id, name, device_code, vendor, device_type, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (site_id, device_name[:100], device_code, '光迅', device_model, 'active', datetime.now()))
            device_id = cursor.lastrowid
            devices_imported += 1
            
            print(f"  导入设备: {device_name[:50]}")
            
            # 读取8个槽位
            for slot_offset in range(8):
                slot_row = row + slot_offset
                if slot_row > ws.max_row:
                    break
                
                slot_num_cell = ws.cell(row=slot_row, column=3)
                card_type_cell = ws.cell(row=slot_row, column=4)
                port_info_cell = ws.cell(row=slot_row, column=5)
                
                slot_number = slot_num_cell.value
                card_type = card_type_cell.value
                port_info = port_info_cell.value
                
                if slot_number and str(slot_number).isdigit():
                    slot_number = int(slot_number)
                    wavelength = parse_wavelength(port_info)
                    direction = parse_direction(port_info)
                    status = 'empty' if not card_type or card_type == '空' else 'active'
                    
                    description = None
                    if port_info and port_info != '空':
                        parts = []
                        if wavelength:
                            parts.append(f"波长:{wavelength}")
                        if direction:
                            parts.append(f"方向:{direction}")
                        parts.append(f"详情:{port_info}")
                        description = " ".join(parts)
                    
                    cursor.execute('''
                        INSERT INTO cards (device_id, slot_number, card_type, status, description)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (device_id, slot_number, card_type, status, description))
                    cards_imported += 1
            
            row += 8  # 跳过已处理的8行
        else:
            row += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"\n=== 导入完成 ===")
    print(f"设备导入: {devices_imported}")
    print(f"板卡导入: {cards_imported}")

def show_summary():
    """显示导入的数据摘要"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("\n=== 云立方机房设备概览 ===")
    
    # 获取站点ID
    cursor.execute('SELECT id FROM sites WHERE name = ?', ('云立方机房',))
    site = cursor.fetchone()
    if not site:
        print("未找到云立方机房数据")
        return
    site_id = site[0]
    
    # 按厂商统计
    cursor.execute('''
        SELECT vendor, COUNT(*) as device_count 
        FROM devices 
        WHERE site_id = ?
        GROUP BY vendor
    ''', (site_id,))
    
    for row in cursor.fetchall():
        vendor, count = row
        print(f"\n【{vendor}】设备数量: {count}")
        
        cursor.execute('''
            SELECT name, device_type, 
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id) as slot_count,
                   (SELECT COUNT(*) FROM cards WHERE device_id = d.id AND status = 'active') as active_count
            FROM devices d
            WHERE site_id = ? AND vendor = ?
        ''', (site_id, vendor))
        
        for device in cursor.fetchall():
            name, dtype, total, active = device
            print(f"  - {name[:45]:<45} {active}/{total} 槽位在用")
    
    # 波长统计
    print("\n=== 波长使用情况 ===")
    cursor.execute('''
        SELECT c.description
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ? AND c.status = 'active' AND c.description LIKE '%波长:%'
    ''', (site_id,))
    
    wavelengths = {}
    for row in cursor.fetchall():
        desc = row[0]
        if desc:
            match = re.search(r'波长:(\d{3}\.\d)', desc)
            if match:
                wl = match.group(1)
                wavelengths[wl] = wavelengths.get(wl, 0) + 1
    
    sorted_wl = sorted(wavelengths.items(), key=lambda x: float(x[0]))
    print(f"在用波长数量: {len(sorted_wl)}")
    if sorted_wl:
        print("波长分布: ", ", ".join([f"{w[0]}nm({w[1]})" for w in sorted_wl]))
    
    # 统计空置槽位
    cursor.execute('''
        SELECT COUNT(*) 
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ? AND c.status = 'empty'
    ''', (site_id,))
    empty_count = cursor.fetchone()[0]
    
    cursor.execute('''
        SELECT COUNT(*) 
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        WHERE d.site_id = ?
    ''', (site_id,))
    total_count = cursor.fetchone()[0]
    
    print(f"\n槽位利用率: {total_count - empty_count}/{total_count} ({(total_count - empty_count) / total_count * 100:.1f}%)")
    
    conn.close()

if __name__ == '__main__':
    print("=" * 60)
    print("云立方机房业务板数据导入工具")
    print("=" * 60)
    print(f"数据库: {DB_PATH}")
    print(f"Excel文件: {EXCEL_PATH}")
    
    import_yunlifang()
    show_summary()
    
    print("\n" + "=" * 60)
    print("✅ 数据导入完成！")
    print("=" * 60)
