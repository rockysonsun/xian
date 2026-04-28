#!/usr/bin/env python3
"""
导入嘉定机房业务板使用情况 - V2版本
使用与云立方相同的逻辑：
1. 正确处理 N2MD02 的 IN1OUT1/IN2OUT2 双波长
2. 完整保留原始"使用状况"作为 description
"""

import sqlite3
import openpyxl
from datetime import datetime
import re

DB_PATH = '/Users/rocky/.openclaw/workspace/transmission_resource.db'
EXCEL_PATH = '/Users/rocky/.openclaw/workspace/2026传输资源分配表.xlsx'


def get_or_create_site(cursor, site_name):
    """获取或创建站点"""
    cursor.execute('SELECT id FROM sites WHERE name = ?', (site_name,))
    site = cursor.fetchone()
    if site:
        return site[0]
    else:
        cursor.execute('INSERT INTO sites (name, code, location) VALUES (?, ?, ?)', 
                      (site_name, site_name[:3].upper(), site_name))
        return cursor.lastrowid


def parse_n2md02_wavelengths(usage_text):
    """
    专门解析 N2MD02 的双波长格式
    支持格式：
    - IN1/OUT1(方向波长)/IN2/OUT2(方向波长)
    - IN1OUT1-xxx-波长/IN2OUT2-xxx-波长
    """
    if not usage_text or usage_text == '空':
        return []
    
    text = str(usage_text)
    results = []
    
    # 模式1: IN1/OUT1(方向波长)/IN2/OUT2(方向波长)
    pattern1 = r'IN1[/]?OUT1\(([^)]+)\)(?:.*?IN2[/]?OUT2\(([^)]+)\))?'
    match1 = re.search(pattern1, text)
    if match1:
        # IN1OUT1
        in1_content = match1.group(1)
        wl_match = re.search(r'(\d{3}\.\d)', in1_content)
        dir_match = re.search(r'(.+?方向)', in1_content)
        if wl_match:
            results.append({
                'port': 'IN1OUT1',
                'wavelength': wl_match.group(1),
                'direction': dir_match.group(1) if dir_match else None
            })
        
        # IN2OUT2
        if match1.group(2):
            in2_content = match1.group(2)
            wl_match = re.search(r'(\d{3}\.\d)', in2_content)
            dir_match = re.search(r'(.+?方向)', in2_content)
            if wl_match:
                results.append({
                    'port': 'IN2OUT2',
                    'wavelength': wl_match.group(1),
                    'direction': dir_match.group(1) if dir_match else None
                })
        return results
    
    # 模式2: IN1OUT1-xxx-波长/IN2OUT2-xxx-波长
    in1_pattern = r'IN1[/]?OUT1[^\d]*(\d{3}\.\d)'
    in2_pattern = r'IN2[/]?OUT2[^\d]*(\d{3}\.\d)'
    
    in1_match = re.search(in1_pattern, text)
    in2_match = re.search(in2_pattern, text)
    
    # 提取方向
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '外高桥' in text:
        directions.append('外高桥')
    if '苏州' in text:
        directions.append('苏州')
    direction = '-'.join(directions) if directions else None
    
    if in1_match:
        results.append({
            'port': 'IN1OUT1',
            'wavelength': in1_match.group(1),
            'direction': direction
        })
    
    if in2_match:
        results.append({
            'port': 'IN2OUT2',
            'wavelength': in2_match.group(1),
            'direction': direction
        })
    
    return results


def parse_single_wavelength(usage_text):
    """解析单波长板卡的波长"""
    if not usage_text or usage_text == '空':
        return None
    
    text = str(usage_text)
    
    # 提取波长
    wl_match = re.search(r'(\d{3}\.\d)', text)
    if not wl_match:
        return None
    
    # 提取方向
    directions = []
    if '云立方' in text:
        directions.append('云立方')
    if '常熟' in text:
        directions.append('常熟')
    if '嘉定' in text:
        directions.append('嘉定')
    if '吴江' in text:
        directions.append('吴江')
    if '外高桥' in text:
        directions.append('外高桥')
    if '苏州' in text:
        directions.append('苏州')
    
    return {
        'port': 'IN/OUT',
        'wavelength': wl_match.group(1),
        'direction': '-'.join(directions) if directions else None
    }


def import_jiading():
    """导入嘉定机房数据"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 打开Excel文件
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['嘉定机房业务板使用情况']
    
    # 获取或创建嘉定站点
    site_id = get_or_create_site(cursor, '嘉定机房')
    print(f"站点ID: {site_id}")
    
    devices_imported = 0
    cards_imported = 0
    ports_imported = 0
    
    current_device_id = None
    current_vendor = None
    current_model = None
    
    # 从第4行开始读取数据 (第1行标题，第2-3行表头)
    for row in range(4, ws.max_row + 1):
        # 读取各列数据
        vendor_cell = ws.cell(row=row, column=1).value  # A列: 品牌
        model_cell = ws.cell(row=row, column=2).value   # B列: 设备型号
        device_code_cell = ws.cell(row=row, column=3).value  # C列: 设备编号
        slot_cell = ws.cell(row=row, column=4).value    # D列: 槽位
        card_type_cell = ws.cell(row=row, column=5).value  # E列: 板卡型号
        usage_cell = ws.cell(row=row, column=6).value   # F列: 使用状况
        
        # 跳过表头行
        if device_code_cell == '设备编号':
            continue
        
        # 如果有设备编号，说明是新设备
        if device_code_cell and str(device_code_cell).strip() and str(device_code_cell).strip() != 'None':
            device_code = str(device_code_cell).strip()
            
            # 确定厂商
            if vendor_cell and str(vendor_cell).strip():
                current_vendor = str(vendor_cell).strip()
            
            # 确定设备型号
            if model_cell and str(model_cell).strip():
                current_model = str(model_cell).strip()
            else:
                current_model = '未知'
            
            # 创建设备记录
            cursor.execute('''
                INSERT INTO devices (site_id, name, device_code, vendor, device_type, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (site_id, device_code, device_code, current_vendor or '未知', current_model, 'active', datetime.now()))
            current_device_id = cursor.lastrowid
            devices_imported += 1
            
            print(f"  导入设备: {device_code} ({current_vendor} {current_model})")
        
        # 如果有槽位号，导入板卡
        if slot_cell and str(slot_cell).strip().isdigit() and current_device_id:
            slot_number = int(str(slot_cell).strip())
            card_type = str(card_type_cell).strip() if card_type_cell else None
            usage = str(usage_cell).strip() if usage_cell else None
            
            if card_type and card_type != '空':
                # 完整保留原始使用状况作为 description
                description = usage if usage and usage != '空' else None
                
                # 解析波长
                wavelengths = []
                if card_type.upper() == 'N2MD02':
                    wavelengths = parse_n2md02_wavelengths(usage)
                else:
                    wl = parse_single_wavelength(usage)
                    if wl:
                        wavelengths = [wl]
                
                status = 'empty' if not usage or usage == '空' else 'active'
                
                # 创建板卡记录
                cursor.execute('''
                    INSERT INTO cards (device_id, slot_number, card_type, status, description)
                    VALUES (?, ?, ?, ?, ?)
                ''', (current_device_id, slot_number, card_type, status, description))
                card_id = cursor.lastrowid
                cards_imported += 1
                
                # 创建端口记录
                for wl_info in wavelengths:
                    cursor.execute('''
                        INSERT INTO ports (card_id, port_number, port_type, status, wavelength, description)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (card_id, wl_info['port'], '光口', 'active', float(wl_info['wavelength']), wl_info.get('direction')))
                    ports_imported += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"\n=== 导入完成 ===")
    print(f"设备导入: {devices_imported}")
    print(f"板卡导入: {cards_imported}")
    print(f"波长端口导入: {ports_imported}")


def verify_data():
    """验证导入的数据"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("\n=== 嘉定机房数据验证 ===")
    
    # 统计
    cursor.execute('''
        SELECT 
            COUNT(DISTINCT d.id) as device_count,
            COUNT(c.id) as card_count,
            COUNT(p.id) as port_count
        FROM devices d
        LEFT JOIN cards c ON d.id = c.device_id
        LEFT JOIN ports p ON c.id = p.card_id
        WHERE d.site_id = (SELECT id FROM sites WHERE name = '嘉定机房')
    ''')
    row = cursor.fetchone()
    print(f"\n设备数: {row[0]}, 板卡数: {row[1]}, 波长端口数: {row[2]}")
    
    # N2MD02 验证
    print("\n=== N2MD02 双波长板卡 ===")
    cursor.execute('''
        SELECT 
            d.name as device_name,
            c.slot_number,
            c.card_type,
            c.description as original_usage,
            GROUP_CONCAT(p.port_number || ':' || p.wavelength, ' | ') as parsed_wavelengths
        FROM cards c
        JOIN devices d ON c.device_id = d.id
        LEFT JOIN ports p ON c.id = p.card_id
        WHERE d.site_id = (SELECT id FROM sites WHERE name = '嘉定机房')
          AND c.card_type = 'N2MD02'
        GROUP BY d.name, c.slot_number
        ORDER BY d.name, c.slot_number
    ''')
    
    for row in cursor.fetchall():
        device, slot, card, usage, wl = row
        print(f"\n{device} - 槽位{slot}:")
        print(f"  原始: {usage}")
        print(f"  解析: {wl}")
    
    conn.close()


if __name__ == '__main__':
    print("=" * 70)
    print("嘉定机房业务板数据导入工具 (V2 - 同云立方逻辑)")
    print("=" * 70)
    
    import_jiading()
    verify_data()
    
    print("\n" + "=" * 70)
    print("✅ 嘉定机房数据导入完成！")
    print("=" * 70)
